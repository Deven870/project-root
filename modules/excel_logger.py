"""
FIX 11: Auto-append signals to Excel paper trade log
Automatically logs trade signals to the Excel tracking workbook.
"""

import os
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

EXCEL_PATH = os.getenv("EXCEL_TRACKER_PATH", "Digitrader_PaperTrading.xlsx")


def log_trade_signal(signal: dict):
    """
    Appends a new row to the Trade Log sheet in the Excel workbook.
    
    Parameters
    ----------
    signal : dict
        Signal dict must contain:
        - symbol
        - current_price
        - predicted_price
        - trend (Bullish/Bearish)
        - confidence (0-1)
        - sentiment (dict with positive/negative keys)
        - stop_loss
        - predicted_return_pct
    """
    try:
        from openpyxl import load_workbook
        
        if not os.path.exists(EXCEL_PATH):
            print(f"Warning: Excel file not found at {EXCEL_PATH}")
            return
        
        wb = load_workbook(EXCEL_PATH)
        
        # Try to get Trade Log sheet, use first sheet if it doesn't exist
        if "📋 Trade Log" in wb.sheetnames:
            ws = wb["📋 Trade Log"]
        elif "Trade Log" in wb.sheetnames:
            ws = wb["Trade Log"]
        else:
            ws = wb.active
        
        # Find next available row
        next_row = ws.max_row + 1
        
        # Fill in columns (adjust column letters based on your Excel structure)
        # Common structure: Date, Symbol, Entry Price, Target, Entry Time, Stop Loss, Status, etc.
        
        ws.cell(next_row, 1).value = datetime.now().strftime("%d-%b-%Y")  # Date
        ws.cell(next_row, 2).value = signal.get("symbol", "")  # Symbol
        ws.cell(next_row, 3).value = "BUY" if "bull" in signal.get("trend", "").lower() else "SELL"  # Signal
        ws.cell(next_row, 4).value = signal.get("current_price", 0)  # Current Price
        ws.cell(next_row, 5).value = signal.get("predicted_price", 0)  # Predicted Price
        ws.cell(next_row, 6).value = signal.get("predicted_return_pct", 0)  # Expected Return %
        ws.cell(next_row, 7).value = signal.get("stop_loss", 0)  # Stop Loss
        ws.cell(next_row, 8).value = signal.get("confidence", 0) * 100  # Confidence %
        
        # Sentiment columns
        sentiment = signal.get("sentiment", {})
        ws.cell(next_row, 9).value = sentiment.get("positive", 0)  # Positive Sentiment
        ws.cell(next_row, 10).value = sentiment.get("negative", 0)  # Negative Sentiment
        
        ws.cell(next_row, 11).value = "OPEN"  # Status
        ws.cell(next_row, 12).value = "Auto-logged by Digitrader"  # Notes
        
        wb.save(EXCEL_PATH)
        print(f"✓ Signal logged for {signal.get('symbol')} in Excel")
    
    except Exception as e:
        print(f"Error logging trade signal to Excel: {e}")


def update_trade_status(symbol, status, pnl=0, exit_price=0, exit_reason=""):
    """
    Update the status of an open trade.
    
    Parameters
    ----------
    symbol : str
        Stock symbol
    status : str
        Trade status (e.g., "CLOSED", "HIT_TARGET", "STOP_HIT")
    pnl : float
        Profit/Loss amount
    exit_price : float
        Exit price achieved
    exit_reason : str
        Reason for exit
    """
    try:
        from openpyxl import load_workbook
        
        if not os.path.exists(EXCEL_PATH):
            return
        
        wb = load_workbook(EXCEL_PATH)
        
        if "📋 Trade Log" in wb.sheetnames:
            ws = wb["📋 Trade Log"]
        elif "Trade Log" in wb.sheetnames:
            ws = wb["Trade Log"]
        else:
            ws = wb.active
        
        # Find the most recent row with this symbol
        for row in range(ws.max_row, 1, -1):
            if ws.cell(row, 2).value == symbol and ws.cell(row, 11).value == "OPEN":
                ws.cell(row, 11).value = status
                ws.cell(row, 13).value = exit_price
                ws.cell(row, 14).value = pnl
                ws.cell(row, 15).value = exit_reason
                ws.cell(row, 16).value = datetime.now().strftime("%d-%b-%Y %H:%M")
                break
        
        wb.save(EXCEL_PATH)
        print(f"✓ Trade status updated for {symbol}")
    
    except Exception as e:
        print(f"Error updating trade status: {e}")
